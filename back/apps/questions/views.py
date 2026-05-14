from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.core.files.uploadedfile import InMemoryUploadedFile
import json
import csv
import io
from openpyxl import load_workbook
from .models import Question, Category
from .serializers import QuestionSerializer, QuestionCreateSerializer, CategorySerializer


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated]


class QuestionViewSet(viewsets.ModelViewSet):
    serializer_class = QuestionSerializer
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return QuestionCreateSerializer
        return QuestionSerializer

    def get_queryset(self):
        queryset = Question.objects.filter(user=self.request.user)
        question_type = self.request.query_params.get('type')
        difficulty = self.request.query_params.get('difficulty')

        if question_type:
            queryset = queryset.filter(type=question_type)
        if difficulty:
            queryset = queryset.filter(difficulty=difficulty)

        return queryset

    def perform_create(self, serializer):
        instance = serializer.save()
        instance.user = self.request.user
        instance.save(update_fields=['user'])

    @action(detail=False, methods=['get'])
    def random(self, request):
        count = int(request.query_params.get('count', 10))
        questions = Question.objects.filter(user=request.user).order_by('?')[:count]
        serializer = self.get_serializer(questions, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def batch_delete(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'error': '请提供要删除的题目ID列表'}, status=status.HTTP_400_BAD_REQUEST)

        deleted_count, _ = Question.objects.filter(user=request.user, id__in=ids).delete()
        return Response({'message': f'成功删除 {deleted_count} 道题目'})

    @action(detail=False, methods=['post'])
    def import_questions(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': '请上传文件'}, status=status.HTTP_400_BAD_REQUEST)

        file_name = file.name.lower()
        try:
            if file_name.endswith('.json'):
                questions = self._parse_json(file)
            elif file_name.endswith('.csv'):
                questions = self._parse_csv(file)
            elif file_name.endswith(('.xlsx', '.xls')):
                questions = self._parse_excel(file)
            else:
                return Response({'error': '不支持的文件格式，请使用JSON、CSV或Excel'}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({'error': f'文件解析失败: {str(e)}'}, status=status.HTTP_400_BAD_REQUEST)

        if not questions:
            return Response({'error': '文件中没有找到有效题目'}, status=status.HTTP_400_BAD_REQUEST)

        created_count = 0
        for q_data in questions:
            try:
                Question.objects.create(
                    user=request.user,
                    type=q_data.get('type', 'choice'),
                    difficulty=q_data.get('difficulty', 'medium'),
                    content=q_data.get('content', ''),
                    options=q_data.get('options', {}),
                    answer=q_data.get('answer', ''),
                    explanation=q_data.get('explanation', ''),
                    created_by_ai=False
                )
                created_count += 1
            except Exception:
                continue

        return Response({'message': f'成功导入 {created_count} 道题目'})

    def _parse_json(self, file):
        content = file.read().decode('utf-8')
        data = json.loads(content)
        if isinstance(data, list):
            return data
        if isinstance(data, dict) and 'questions' in data:
            return data['questions']
        return []

    def _parse_csv(self, file):
        content = file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(content))
        questions = []
        for row in reader:
            options = {}
            if row.get('option_a'):
                options['A'] = row['option_a']
            if row.get('option_b'):
                options['B'] = row['option_b']
            if row.get('option_c'):
                options['C'] = row['option_c']
            if row.get('option_d'):
                options['D'] = row['option_d']

            questions.append({
                'type': row.get('type', 'choice'),
                'difficulty': row.get('difficulty', 'medium'),
                'content': row.get('content', ''),
                'options': options,
                'answer': row.get('answer', ''),
                'explanation': row.get('explanation', '')
            })
        return questions

    def _parse_excel(self, file):
        wb = load_workbook(file, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []

        headers = [str(h).lower() if h else '' for h in rows[0]]
        questions = []
        for row in rows[1:]:
            data = dict(zip(headers, row))
            options = {}
            if data.get('option_a'):
                options['A'] = str(data['option_a'])
            if data.get('option_b'):
                options['B'] = str(data['option_b'])
            if data.get('option_c'):
                options['C'] = str(data['option_c'])
            if data.get('option_d'):
                options['D'] = str(data['option_d'])

            questions.append({
                'type': str(data.get('type', 'choice')),
                'difficulty': str(data.get('difficulty', 'medium')),
                'content': str(data.get('content', '')),
                'options': options,
                'answer': str(data.get('answer', '')),
                'explanation': str(data.get('explanation', ''))
            })
        return questions
